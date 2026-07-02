import re
from typing import Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from backend.models import Tag, Transaction
from backend.services.csv_service import CSVService


class ExportService:
    def __init__(self, db: Session):
        self.db = db

    def export_excel(
        self,
        transactions: list[Transaction],
        file_path: str,
        session_name: str = "Audit",
        client_name_to_code: Optional[Dict[str, str]] = None,
    ) -> str:
        """Export transactions to a formatted multi-sheet Excel workbook."""
        wb = Workbook()

        all_tags = self._tags_by_transaction(transactions)
        suspicious_transactions = self._filter_by_tag(
            transactions, "suspicious", all_tags
        )
        suspicious_breakdown = self._suspicious_breakdown(
            suspicious_transactions, all_tags
        )
        sheets = [
            ("Account Transactions", transactions),
            ("Client", self._filter_by_tag(transactions, "client", all_tags)),
            ("Broker", self._filter_by_tag(transactions, "broker", all_tags)),
            ("Suspicious", suspicious_transactions),
            ("Suspicious - Recurring", suspicious_breakdown["recurring"]),
            ("Suspicious - High Value", suspicious_breakdown["high_value"]),
            ("Suspicious - Other", suspicious_breakdown["other"]),
        ]

        for index, (title, sheet_transactions) in enumerate(sheets):
            ws = wb.active if index == 0 else wb.create_sheet()
            ws.title = title
            self._write_transaction_sheet(
                ws,
                sheet_transactions,
                all_tags,
                title,
                session_name,
                client_name_to_code if title == "Client" else None,
            )
            ws.sheet_properties.tabColor = {
                "Account Transactions": "366092",
                "Client": "70AD47",
                "Broker": "FFC000",
                "Suspicious": "C00000",
                "Suspicious - Recurring": "C00000",
                "Suspicious - High Value": "C00000",
                "Suspicious - Other": "C00000",
            }.get(title, "366092")

        wb.save(file_path)
        return file_path

    def _write_transaction_sheet(
        self,
        ws: Worksheet,
        transactions: list[Transaction],
        tags_by_transaction: dict[int, list[Tag]],
        title: str,
        session_name: str,
        client_name_to_code: Optional[Dict[str, str]] = None,
    ) -> None:
        headers = [
            "ID",
            "Date",
            "Debit",
            "Credit",
            "Description",
            "Party Name",
        ]
        if client_name_to_code is not None:
            headers.append("Client Code")
        headers.extend([
            "Payment Method",
            "Tags",
            "Tag Reasons",
            "Notes",
            "PDF File",
            "Page",
        ])
        header_font = Font(bold=True)
        subtle_side = Side(style="thin", color="D9E2EC")
        thin_border = Border(
            left=subtle_side, right=subtle_side, top=subtle_side, bottom=subtle_side
        )

        header_row = 1

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        sorted_transactions = self._sort_for_sheet(
            transactions, tags_by_transaction, title
        )

        for row, tx in enumerate(sorted_transactions, header_row + 1):
            tags = tags_by_transaction.get(tx.id, [])
            tag_types = [t.tag_type for t in tags]
            tag_reasons = [t.reason for t in tags]

            export_party = self._export_party_name(tags, tx.party_name)

            client_code = ""
            if client_name_to_code is not None:
                for t in tags:
                    if t.tag_type == "client" and t.reason:
                        matched = self._matched_client_name_from_reason(t.reason)
                        if matched:
                            key = CSVService.normalize_client_name_key(matched)
                            client_code = client_name_to_code.get(key, "")
                        break
                if client_code:
                    client_code = CSVService.normalize_client_code(client_code)

            values = [
                tx.id,
                tx.date,
                -tx.amount if tx.amount and tx.amount < 0 else "",
                tx.amount if tx.amount and tx.amount > 0 else "",
                tx.description,
                export_party,
            ]
            if client_name_to_code is not None:
                values.append(client_code)
            values.extend([
                tx.payment_method,
                ", ".join(tag_types),
                "; ".join(reason for reason in tag_reasons if reason),
                tx.user_notes,
                tx.pdf_filename,
                tx.page_number,
            ])

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                desc_col = 5
                wrap_cols = {desc_col, desc_col + 4, desc_col + 5}
                if client_name_to_code is not None:
                    wrap_cols = {desc_col, desc_col + 5, desc_col + 6}
                cell.alignment = Alignment(vertical="top", wrap_text=col in wrap_cols)
                cell.border = thin_border

        if not sorted_transactions:
            empty_row = header_row + 1
            ws.cell(row=empty_row, column=1, value="No transactions in this category")
            ws.cell(row=empty_row, column=1).font = Font(italic=True, color="6B7280")
            ws.cell(row=empty_row, column=1).alignment = Alignment(vertical="top")
            ws.cell(row=empty_row, column=1).border = thin_border
            ws.merge_cells(
                start_row=empty_row,
                start_column=1,
                end_row=empty_row,
                end_column=len(headers),
            )

        data_end_row = ws.max_row
        meta_start_row = ws.max_row + 2
        ws.cell(row=meta_start_row, column=1, value="Sheet Metadata").font = Font(
            bold=True
        )
        ws.merge_cells(
            start_row=meta_start_row,
            start_column=1,
            end_row=meta_start_row,
            end_column=len(headers),
        )

        ws.cell(
            row=meta_start_row + 1, column=1, value="Account / Session"
        ).font = Font(bold=True)
        ws.cell(row=meta_start_row + 1, column=2, value=session_name)

        ws.cell(
            row=meta_start_row + 2, column=1, value="Transaction Count"
        ).font = Font(bold=True)
        ws.cell(row=meta_start_row + 2, column=2, value=len(transactions))

        ws.cell(row=meta_start_row + 3, column=1, value="Sheet Category").font = Font(
            bold=True
        )
        ws.cell(row=meta_start_row + 3, column=2, value=title)

        for r in range(meta_start_row, meta_start_row + 4):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = thin_border

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = (
            f"A{header_row}:{ws.cell(row=data_end_row, column=len(headers)).coordinate}"
        )

        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        print(f"[ExportService] Error calculating column width: {e}")
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        for column in ("C", "D"):
            for cell in ws[column][header_row:]:
                cell.number_format = "#,##0.00"

    def _export_party_name(
        self, tags: list[Tag], fallback: Optional[str]
    ) -> Optional[str]:
        if not tags:
            return fallback
        for t in tags:
            if not t.reason:
                continue
            if t.tag_type == "client":
                name = self._matched_client_name_from_reason(t.reason)
                if name:
                    return name
            elif t.tag_type == "broker":
                name = self._matched_broker_name_from_reason(t.reason)
                if name:
                    return name
        for t in tags:
            if t.tag_type == "suspicious" and t.reason:
                name = self._party_from_suspicious_reason(t.reason)
                if name:
                    return name
        if any(t.tag_type == "suspicious" for t in tags):
            return ""
        return fallback

    def _matched_client_name_from_reason(self, reason: str) -> Optional[str]:
        m = re.search(r"Fuzzy match:\s*'([^']+)'", reason)
        if m:
            return m.group(1).strip()
        m = re.search(r"Phone match:\s*\d+\s*->\s*'?([^'\[]+?)'?", reason)
        if m:
            return m.group(1).strip().strip("'\"")
        return None

    def _matched_broker_name_from_reason(self, reason: str) -> Optional[str]:
        m = re.search(r"Broker match:\s*'([^']+)'", reason)
        if m:
            return m.group(1).strip()
        return None

    _SUSPICIOUS_PARTY_SKIP = frozenset(
        {"mr", "mr.", "ms", "ms.", "mrs", "mrs.", "dr", "dr.", "same party"}
    )

    def _party_from_suspicious_reason(self, reason: str) -> Optional[str]:
        m = re.search(r"\bwith\s+([^:]+):", reason, re.IGNORECASE)
        if not m:
            return None
        party = re.sub(r"\s+", " ", m.group(1).strip())
        if not party or party.lower() in self._SUSPICIOUS_PARTY_SKIP:
            return None
        # ponytail: skip 1–4 char UPI fragments (Gene, LALI, hars)
        if len(party) <= 4 and party.isalpha():
            return None
        return party

    def _tags_by_transaction(
        self, transactions: list[Transaction]
    ) -> dict[int, list[Tag]]:
        return {tx.id: list(tx.tags) for tx in transactions if tx.id is not None}

    def _filter_by_tag(
        self,
        transactions: list[Transaction],
        tag_type: str,
        tags_by_transaction: dict[int, list[Tag]],
    ) -> list[Transaction]:
        return [
            tx
            for tx in transactions
            if any(
                tag.tag_type == tag_type for tag in tags_by_transaction.get(tx.id, [])
            )
        ]

    def _suspicious_breakdown(
        self,
        transactions: list[Transaction],
        tags_by_transaction: dict[int, list[Tag]],
    ) -> dict[str, list[Transaction]]:
        grouped = {"recurring": [], "high_value": [], "other": []}
        for tx in transactions:
            grouped[self._suspicious_category(tx, tags_by_transaction)].append(tx)
        return grouped

    def _suspicious_category(
        self, tx: Transaction, tags_by_transaction: dict[int, list[Tag]]
    ) -> str:
        suspicious_tags = [
            tag
            for tag in tags_by_transaction.get(tx.id, [])
            if tag.tag_type == "suspicious"
        ]
        reason = " ".join(tag.reason or "" for tag in suspicious_tags).lower()
        if "recurring" in reason:
            return "recurring"
        if "exceeds threshold" in reason:
            return "high_value"
        return "other"

    def _sort_for_sheet(
        self,
        transactions: list[Transaction],
        tags_by_transaction: dict[int, list[Tag]],
        title: str,
    ) -> list[Transaction]:
        if not title.startswith("Suspicious - Recurring"):
            return transactions
        return sorted(
            transactions,
            key=lambda tx: (
                self._party_sort_key(tx),
                tx.date or "",
                tx.id or 0,
            ),
        )

    def _party_sort_key(self, tx: Transaction) -> str:
        value = tx.party_name or tx.description or tx.raw_text or "Unknown"
        return re.sub(r"\s+", " ", value).strip().casefold()
