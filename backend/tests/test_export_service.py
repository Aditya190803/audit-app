import os
import tempfile
import unittest
from unittest.mock import patch

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import backend.api.routes.export as export_route
from backend.database import Base
from backend.models import AuditSession, Tag, Transaction
from backend.security import export_path_token
from backend.services.export_service import ExportService


class ExportServiceTests(unittest.TestCase):
    def test_normalize_client_code_strips_leading_quote(self):
        from backend.services.csv_service import CSVService
        self.assertEqual(CSVService.normalize_client_code("'S2052"), "S2052")
        self.assertEqual(CSVService.normalize_client_code("'CE01943"), "CE01943")
        self.assertEqual(CSVService.normalize_client_code("CE01943"), "CE01943")

    def test_export_party_name_from_client_tag_not_parser_party(self):
        svc = ExportService(None)
        tags = [
            Tag(
                transaction_id=1,
                tag_type="client",
                reason="Fuzzy match: 'BHAVINI C SHAH' (score: 1.0)",
            )
        ]
        self.assertEqual(
            svc._export_party_name(tags, "ICIC"),
            "BHAVINI C SHAH",
        )

    def test_export_route_fills_client_code_from_session_csv(self):
        import csv

        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = os.path.join(tmpdir, "clients.csv")
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["CLIENTCODE", "CLIENTNAME"])
                w.writerow(["'CE01943", "BHAVINI C SHAH"])

            output_xlsx = os.path.join(tmpdir, "exports", "audit.xlsx")

            engine = create_engine("sqlite:///:memory:")
            Base.metadata.create_all(engine)
            db = sessionmaker(bind=engine)()
            try:
                session = AuditSession(
                    name="ICIC",
                    csv_path=csv_path,
                    settings_snapshot={
                        "client_name_column": "CLIENTNAME",
                        "client_code_column": "CLIENTCODE",
                    },
                )
                db.add(session)
                db.commit()

                tx = Transaction(
                    session_id=session.id,
                    date="2025-03-16",
                    amount=-10000,
                    description="IMPS/P2A/.../BHAVINI C SHAH",
                    party_name="ICIC",
                    payment_method="IMPS",
                    page_number=2,
                )
                db.add(tx)
                db.commit()
                db.refresh(tx)
                db.add(
                    Tag(
                        transaction_id=tx.id,
                        tag_type="client",
                        reason="Fuzzy match: 'BHAVINI C SHAH' (score: 1.0)",
                    )
                )
                db.commit()

                with patch.object(export_route, "EXPORT_DIR", os.path.join(tmpdir, "exports")):
                    resp = export_route.export_excel(
                        session.id, file_path="audit.xlsx", transaction_ids=None, db=db
                    )

                wb = load_workbook(resp["file_path"])
                try:
                    ws = wb["Client"]
                    self.assertEqual(ws["G2"].value, "CE01943")
                    self.assertEqual(ws["F2"].value, "BHAVINI C SHAH")
                finally:
                    wb.close()
            finally:
                db.close()

    def test_export_party_name_empty_for_suspicious_junk_reason(self):
        svc = ExportService(None)
        tags = [
            Tag(
                transaction_id=1,
                tag_type="suspicious",
                reason="Recurring debit of ₹300.00 with MR.: 12 matching transactions",
            )
        ]
        self.assertEqual(svc._export_party_name(tags, "MR."), "")

    def test_export_party_name_from_suspicious_reason_when_usable(self):
        svc = ExportService(None)
        tags = [
            Tag(
                transaction_id=1,
                tag_type="suspicious",
                reason=(
                    "Recurring debit of ₹8,022.00 with Loan Recovery For795000021260: "
                    "6 matching transactions"
                ),
            )
        ]
        self.assertEqual(
            svc._export_party_name(tags, "ICIC"),
            "Loan Recovery For795000021260",
        )

    def setUp(self):
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        self.db = sessionmaker(bind=engine)()

    def tearDown(self):
        self.db.close()

    def assert_workbook_has_no_cell_fills(self, wb):
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    self.assertIsNone(
                        cell.fill.fill_type,
                        f"Unexpected fill on {ws.title}!{cell.coordinate}",
                    )

    def test_export_route_rejects_unapproved_absolute_path(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_xlsx = os.path.join(tmpdir, "nested", "audit.xlsx")

            with (
                patch.object(
                    export_route, "EXPORT_DIR", os.path.join(tmpdir, "exports")
                ),
                self.assertRaises(HTTPException),
            ):
                export_route._ensure_export_path(output_xlsx)

    def test_export_route_accepts_save_dialog_approved_absolute_path(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_xlsx = os.path.join(tmpdir, "nested", "audit.xlsx")

            with (
                patch.dict(os.environ, {"AUDIT_EXPORT_PATH_SECRET": "test-secret"}),
                patch.object(
                    export_route, "EXPORT_DIR", os.path.join(tmpdir, "exports")
                ),
            ):
                token = export_path_token(output_xlsx)
                self.assertEqual(
                    export_route._ensure_export_path(output_xlsx, token),
                    os.path.realpath(output_xlsx),
                )
                self.assertTrue(os.path.isdir(os.path.dirname(output_xlsx)))

    def test_excel_export_creates_required_workbook_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_xlsx = os.path.join(tmpdir, "audit.xlsx")

            session = AuditSession(name="AAAC HDFC")
            self.db.add(session)
            self.db.commit()

            client_tx = Transaction(
                session_id=session.id,
                date="2026-01-02",
                amount=-1250,
                description="NEFT TO CLIENT ONE",
                party_name="CLIENT ONE",
                payment_method="NEFT",
                page_number=1,
            )
            broker_tx = Transaction(
                session_id=session.id,
                date="2026-01-03",
                amount=5000,
                description="UPI FROM BROKER TWO",
                party_name="BROKER TWO",
                payment_method="UPI",
                page_number=2,
            )
            suspicious_tx = Transaction(
                session_id=session.id,
                date="2026-01-04",
                amount=-99999,
                description="CASH WITHDRAWAL",
                party_name="UNKNOWN",
                payment_method="CASH",
                page_number=3,
            )
            self.db.add_all([client_tx, broker_tx, suspicious_tx])
            self.db.commit()
            for tx in (client_tx, broker_tx, suspicious_tx):
                self.db.refresh(tx)

            self.db.add_all(
                [
                    Tag(
                        transaction_id=client_tx.id,
                        tag_type="client",
                        reason="Fuzzy match: 'CLIENT ONE' (score: 0.95)",
                    ),
                    Tag(
                        transaction_id=broker_tx.id,
                        tag_type="broker",
                        reason="Matched broker list",
                    ),
                    Tag(
                        transaction_id=suspicious_tx.id,
                        tag_type="suspicious",
                        reason="Large cash movement",
                    ),
                ]
            )
            self.db.commit()

            ExportService(self.db).export_excel(
                [client_tx, broker_tx, suspicious_tx],
                output_xlsx,
                session.name,
                client_name_to_code={"CLIENT ONE": "C001"},
            )

            wb = load_workbook(output_xlsx)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    [
                        "Account Transactions",
                        "Client",
                        "Broker",
                        "Suspicious",
                        "Suspicious - Recurring",
                        "Suspicious - High Value",
                        "Suspicious - Other",
                    ],
                )

                self.assert_workbook_has_no_cell_fills(wb)

                ws_acc = wb["Account Transactions"]
                max_r = ws_acc.max_row
                self.assertEqual(ws_acc["A1"].value, "ID")
                self.assertEqual(ws_acc["B1"].value, "Date")
                self.assertEqual(ws_acc["H2"].value, "client")
                self.assertEqual(ws_acc[f"B{max_r - 2}"].value, "AAAC HDFC")
                self.assertEqual(ws_acc[f"B{max_r - 1}"].value, 3)

                ws_client = wb["Client"]
                max_r_client = ws_client.max_row
                self.assertEqual(ws_client["G1"].value, "Client Code")
                self.assertEqual(ws_client["G2"].value, "C001")
                self.assertEqual(ws_client[f"B{max_r_client - 1}"].value, 1)

                ws_broker = wb["Broker"]
                max_r_broker = ws_broker.max_row
                self.assertEqual(ws_broker[f"B{max_r_broker - 1}"].value, 1)

                ws_suspicious = wb["Suspicious"]
                max_r_suspicious = ws_suspicious.max_row
                self.assertEqual(ws_suspicious[f"B{max_r_suspicious - 1}"].value, 1)

                ws_other = wb["Suspicious - Other"]
                max_r_other = ws_other.max_row
                self.assertEqual(ws_other[f"B{max_r_other - 1}"].value, 1)
            finally:
                wb.close()

    def test_excel_export_groups_recurring_suspicious_transactions_by_name(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_xlsx = os.path.join(tmpdir, "audit.xlsx")

            session = AuditSession(name="recurring")
            self.db.add(session)
            self.db.commit()

            tx1 = Transaction(
                session_id=session.id,
                date="2026-01-01",
                amount=-5000,
                party_name="ALPHA CLIENT",
            )
            tx2 = Transaction(
                session_id=session.id,
                date="2026-01-15",
                amount=-5000,
                party_name="ALPHA CLIENT",
            )
            tx3 = Transaction(
                session_id=session.id,
                date="2026-01-08",
                amount=-7000,
                party_name="BETA CLIENT",
            )
            self.db.add_all([tx1, tx2, tx3])
            self.db.commit()
            for tx in (tx1, tx2, tx3):
                self.db.refresh(tx)

            self.db.add_all(
                [
                    Tag(
                        transaction_id=tx3.id,
                        tag_type="suspicious",
                        reason="Recurring transaction to same party",
                    ),
                    Tag(
                        transaction_id=tx1.id,
                        tag_type="suspicious",
                        reason="Recurring transaction to same party",
                    ),
                    Tag(
                        transaction_id=tx2.id,
                        tag_type="suspicious",
                        reason="Recurring transaction to same party",
                    ),
                ]
            )
            self.db.commit()

            ExportService(self.db).export_excel(
                [tx3, tx1, tx2], output_xlsx, session.name
            )

            wb = load_workbook(output_xlsx)
            try:
                ws = wb["Suspicious - Recurring"]
                max_r = ws.max_row
                self.assertEqual(ws[f"B{max_r - 1}"].value, 3)
                self.assertIn(ws["F2"].value, (None, ""))
                self.assertIn(ws["F3"].value, (None, ""))
                self.assertIn(ws["F4"].value, (None, ""))
            finally:
                wb.close()
